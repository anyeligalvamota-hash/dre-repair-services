from fastapi import FastAPI, APIRouter, HTTPException, Depends, UploadFile, File, Form, Body
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict, EmailStr
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timezone, timedelta
import bcrypt
import jwt
import resend
import aiofiles
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import io
from fastapi.responses import StreamingResponse
import requests

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Resend API key
resend.api_key = os.environ.get('RESEND_API_KEY')

# JWT settings
SECRET_KEY = os.environ['JWT_SECRET']
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 43200  # 30 days

# Create the main app
app = FastAPI()
api_router = APIRouter(prefix="/api")
security = HTTPBearer()

# File upload directory
UPLOAD_DIR = Path("/app/uploads")
UPLOAD_DIR.mkdir(exist_ok=True)

# ==================== MODELS ====================

class UserRole(BaseModel):
    visitante: bool = False
    colaborador: bool = False
    supervisor: bool = False
    gerente: bool = False
    director: bool = False
    ceo: bool = False
    compras: bool = False
    admin: bool = False

class User(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    email: EmailStr
    nombre: str
    departamento: Optional[str] = None
    roles: UserRole = Field(default_factory=UserRole)
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class UserCreate(BaseModel):
    email: EmailStr
    password: str
    nombre: str
    departamento: Optional[str] = None
    roles: Optional[UserRole] = None

class UserLogin(BaseModel):
    email: EmailStr
    password: str

class Requisicion(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: f"REQ-{str(uuid.uuid4())[:8].upper()}")
    tipo: str  # "compra", "materia_prima", "servicios"
    numero_solicitud: str = Field(default_factory=lambda: f"REQ-{str(uuid.uuid4())[:8].upper()}")
    fecha: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    solicitante: str
    solicitante_id: Optional[str] = None
    departamento: str
    prioridad: str  # "Normal", "Alta"
    fecha_requerida: Optional[datetime] = None
    
    # Campos comunes
    material_descripcion: Optional[str] = None
    cantidad: Optional[float] = None
    unidad_medida: Optional[str] = None
    comentarios: str
    
    # Campos específicos de materia prima
    requerido_por: Optional[str] = None
    area: Optional[str] = None
    proyecto: Optional[str] = None
    proceso: Optional[str] = None
    main_part: Optional[str] = None
    part_number: Optional[str] = None
    modelo: Optional[str] = None
    descripcion: Optional[str] = None
    color: Optional[str] = None
    cantidad_solicitada: Optional[float] = None
    calidad_solicitada: Optional[str] = None
    referencia: Optional[str] = None
    imagen_url: Optional[str] = None
    
    # Campos de aprobación
    supervisor_aprobador: str
    supervisor_aprobador_id: Optional[str] = None
    recibido_por: Optional[str] = None
    aprobado_por: Optional[str] = None
    autorizacion_ehs: Optional[str] = None
    
    # Proveedor
    proveedor_sugerido: Optional[str] = None
    
    # Estado
    estado: str = "pendiente"  # pendiente, cotizaciones_cargadas, enviado_aprobacion, aprobado, rechazado, hold
    
    # Metadata
    cotizaciones_count: int = 0
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    comentario_rechazo: Optional[str] = None

class RequisicionCreate(BaseModel):
    tipo: str
    departamento: str
    prioridad: str
    fecha_requerida: Optional[str] = None
    material_descripcion: Optional[str] = None
    cantidad: Optional[float] = None
    unidad_medida: Optional[str] = None
    comentarios: str
    requerido_por: Optional[str] = None
    area: Optional[str] = None
    proyecto: Optional[str] = None
    proceso: Optional[str] = None
    main_part: Optional[str] = None
    part_number: Optional[str] = None
    modelo: Optional[str] = None
    descripcion: Optional[str] = None
    color: Optional[str] = None
    cantidad_solicitada: Optional[float] = None
    calidad_solicitada: Optional[str] = None
    referencia: Optional[str] = None
    supervisor_aprobador: str
    recibido_por: Optional[str] = None
    autorizacion_ehs: Optional[str] = None
    proveedor_sugerido: Optional[str] = None

class Cotizacion(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    requisicion_id: str
    numero_cotizacion: str
    proveedor: str
    monto_rdp: float
    monto_usd: float
    descripcion: str
    archivo_url: Optional[str] = None
    fecha_carga: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    aprobada: bool = False
    rechazada: bool = False
    comentario_rechazo: Optional[str] = None
    cargado_por: str
    cargado_por_id: str

class CotizacionCreate(BaseModel):
    requisicion_id: str
    numero_cotizacion: str
    proveedor: str
    monto_rdp: float
    monto_usd: float
    descripcion: str

class PurchaseDocument(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: f"PD-{str(uuid.uuid4())[:8].upper()}")
    requisicion_id: str
    cotizacion_id: str
    
    # Campos obligatorios
    part_numbers: str
    supplier: str
    po: str
    po_date: datetime
    month: str
    department: str
    siga: str
    description: str
    uom: str
    qty: float
    unit_price_rdp: float
    itbis_rdp: float
    net_price_rdp: float
    dollar_rate: float
    mpq: Optional[str] = None
    proyecto: Optional[str] = None
    approved_by: str
    
    # Campos calculados automáticamente
    unit_price_usd: float = 0.0
    net_price_usd: float = 0.0
    qty_received: float = 0.0
    qty_pending: float = 0.0
    net_price_qty_pending: float = 0.0
    request_day: str = ""
    request_week: str = ""
    request_quarter: str = ""
    purchasing_response_time: int = 0
    supplier_lead_time: int = 0
    
    # Estados y seguimiento
    order_status: str = "Pendiente pago"  # Pendiente pago, En tránsito, En HOLD, Completado
    loc_int: Optional[str] = None
    closing_cycle: Optional[str] = None
    factura_ncf: Optional[str] = None
    termino_pago: Optional[str] = None
    card: Optional[str] = None
    responsible_closing: str
    
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class PDCreate(BaseModel):
    requisicion_id: str
    cotizacion_id: str
    part_numbers: str
    supplier: str
    po: str
    po_date: str
    department: str
    siga: str
    description: str
    uom: str
    qty: float
    unit_price_rdp: float
    itbis_rdp: float
    dollar_rate: float
    mpq: Optional[str] = None
    proyecto: Optional[str] = None
    approved_by: str
    loc_int: Optional[str] = None
    factura_ncf: Optional[str] = None
    termino_pago: Optional[str] = None
    card: Optional[str] = None

# ==================== AUTH FUNCTIONS ====================

def create_access_token(data: dict):
    to_encode = data.copy()
    expire = datetime.now(timezone.utc) + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

def verify_token(token: str):
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        return payload
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expirado")
    except jwt.JWTError:
        raise HTTPException(status_code=401, detail="Token inválido")

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    token = credentials.credentials
    payload = verify_token(token)
    user_id = payload.get("sub")
    if not user_id:
        raise HTTPException(status_code=401, detail="Token inválido")
    
    user = await db.users.find_one({"id": user_id}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=401, detail="Usuario no encontrado")
    return User(**user)

# ==================== HELPER FUNCTIONS ====================

def calculate_pd_fields(pd_data: Dict) -> Dict:
    """Calcula los campos automáticos del PD"""
    # Conversión USD
    pd_data['unit_price_usd'] = round(pd_data['unit_price_rdp'] / pd_data['dollar_rate'], 2)
    pd_data['net_price_usd'] = round(pd_data['net_price_rdp'] / pd_data['dollar_rate'], 2)
    
    # Cantidades pendientes
    pd_data['qty_pending'] = pd_data['qty'] - pd_data.get('qty_received', 0)
    pd_data['net_price_qty_pending'] = round(pd_data['net_price_rdp'] * (pd_data['qty_pending'] / pd_data['qty']), 2)
    
    # Fechas
    po_date = pd_data['po_date'] if isinstance(pd_data['po_date'], datetime) else datetime.fromisoformat(pd_data['po_date'])
    pd_data['request_day'] = po_date.strftime('%Y-%m-%d')
    pd_data['request_week'] = f"W{po_date.isocalendar()[1]}"
    pd_data['request_quarter'] = f"Q{(po_date.month - 1) // 3 + 1}"
    
    # Tiempos de respuesta (calculados desde la fecha de creación de la requisición)
    pd_data['purchasing_response_time'] = 0  # Se calculará con la fecha de requisición
    pd_data['supplier_lead_time'] = 0
    
    pd_data['month'] = po_date.strftime('%B %Y')
    pd_data['responsible_closing'] = pd_data.get('approved_by', 'N/A')
    
    return pd_data

async def send_email_notification(to_email: str, subject: str, html_content: str):
    """Envía notificación por correo usando Resend"""
    try:
        resend.Emails.send({
            "from": "onboarding@resend.dev",
            "to": to_email,
            "subject": subject,
            "html": html_content
        })
        return True
    except Exception as e:
        logging.error(f"Error enviando correo: {str(e)}")
        return False

async def get_exchange_rate():
    """Obtiene la tasa de cambio actual del dólar"""
    try:
        # API de tasas de cambio (usando exchangerate-api.com como ejemplo)
        response = requests.get("https://api.exchangerate-api.com/v4/latest/USD")
        data = response.json()
        # Tasa DOP (Peso Dominicano)
        rate = data.get('rates', {}).get('DOP', 58.0)
        return round(rate, 2)
    except:
        # Tasa por defecto si falla la API
        return 58.0

# ==================== HEALTH CHECK ====================

@api_router.get("/")
async def root():
    return {"status": "ok", "message": "DRE Repair Services API", "version": "1.0"}

# ==================== AUTH ENDPOINTS ====================

@api_router.post("/auth/register")
async def register(user_data: UserCreate):
    # Verificar si el usuario ya existe
    existing = await db.users.find_one({"email": user_data.email})
    if existing:
        raise HTTPException(status_code=400, detail="El correo ya está registrado")
    
    # Hash de la contraseña
    hashed_password = bcrypt.hashpw(user_data.password.encode('utf-8'), bcrypt.gensalt())
    
    # Crear usuario
    user = User(
        email=user_data.email,
        nombre=user_data.nombre,
        departamento=user_data.departamento,
        roles=user_data.roles if user_data.roles else UserRole(colaborador=True)
    )
    
    user_dict = user.model_dump()
    user_dict['password'] = hashed_password.decode('utf-8')
    user_dict['created_at'] = user_dict['created_at'].isoformat()
    
    await db.users.insert_one(user_dict)
    
    # Crear token
    token = create_access_token({"sub": user.id, "email": user.email})
    
    return {"user": user, "token": token}

@api_router.post("/auth/login")
async def login(credentials: UserLogin):
    user_doc = await db.users.find_one({"email": credentials.email})
    if not user_doc:
        raise HTTPException(status_code=401, detail="Credenciales inválidas")
    
    # Verificar contraseña
    if not bcrypt.checkpw(credentials.password.encode('utf-8'), user_doc['password'].encode('utf-8')):
        raise HTTPException(status_code=401, detail="Credenciales inválidas")
    
    user = User(**{k: v for k, v in user_doc.items() if k != 'password'})
    token = create_access_token({"sub": user.id, "email": user.email})
    
    return {"user": user, "token": token}

@api_router.get("/auth/me")
async def get_me(current_user: User = Depends(get_current_user)):
    return current_user

# ==================== REQUISICION ENDPOINTS ====================

@api_router.post("/requisiciones/public", response_model=Requisicion)
async def create_requisicion_public(req_data: RequisicionCreate):
    """Crear requisición sin autenticación (visitantes)"""
    solicitante = req_data.requerido_por or "Visitante"
    solicitante_id = None
    
    requisicion = Requisicion(
        tipo=req_data.tipo,
        solicitante=solicitante,
        solicitante_id=solicitante_id,
        departamento=req_data.departamento,
        prioridad=req_data.prioridad,
        fecha_requerida=datetime.fromisoformat(req_data.fecha_requerida) if req_data.fecha_requerida else None,
        material_descripcion=req_data.material_descripcion,
        cantidad=req_data.cantidad,
        unidad_medida=req_data.unidad_medida,
        comentarios=req_data.comentarios,
        requerido_por=req_data.requerido_por,
        area=req_data.area,
        proyecto=req_data.proyecto,
        proceso=req_data.proceso,
        main_part=req_data.main_part,
        part_number=req_data.part_number,
        modelo=req_data.modelo,
        descripcion=req_data.descripcion,
        color=req_data.color,
        cantidad_solicitada=req_data.cantidad_solicitada,
        calidad_solicitada=req_data.calidad_solicitada,
        referencia=req_data.referencia,
        supervisor_aprobador=req_data.supervisor_aprobador,
        recibido_por=req_data.recibido_por,
        autorizacion_ehs=req_data.autorizacion_ehs,
        proveedor_sugerido=req_data.proveedor_sugerido
    )
    
    req_dict = requisicion.model_dump()
    req_dict['fecha'] = req_dict['fecha'].isoformat()
    req_dict['created_at'] = req_dict['created_at'].isoformat()
    req_dict['updated_at'] = req_dict['updated_at'].isoformat()
    if req_dict['fecha_requerida']:
        req_dict['fecha_requerida'] = req_dict['fecha_requerida'].isoformat()
    
    await db.requisiciones.insert_one(req_dict)
    
    # Enviar correo de confirmación al email de compras
    email_html = f"""
    <h2>Nueva Requisición Pública - DRE Repair Services</h2>
    <p>Se ha creado una nueva requisición desde el portal público.</p>
    <p><strong>Número de solicitud:</strong> {requisicion.numero_solicitud}</p>
    <p><strong>Tipo:</strong> {requisicion.tipo}</p>
    <p><strong>Solicitante:</strong> {solicitante}</p>
    <p><strong>Departamento:</strong> {requisicion.departamento}</p>
    <p><strong>Estado:</strong> Pendiente de cotizaciones</p>
    """
    
    notification_email = os.environ.get('NOTIFICATION_EMAIL', 'compras-drerepairservices@hotmail.com')
    await send_email_notification(notification_email, 
                                 f"Nueva Requisición {requisicion.numero_solicitud}", 
                                 email_html)
    
    return requisicion

@api_router.post("/requisiciones", response_model=Requisicion)
async def create_requisicion(req_data: RequisicionCreate, current_user: User = Depends(get_current_user)):
    # Usuario autenticado
    solicitante = current_user.nombre
    solicitante_id = current_user.id
    
    requisicion = Requisicion(
        tipo=req_data.tipo,
        solicitante=solicitante,
        solicitante_id=solicitante_id,
        departamento=req_data.departamento,
        prioridad=req_data.prioridad,
        fecha_requerida=datetime.fromisoformat(req_data.fecha_requerida) if req_data.fecha_requerida else None,
        material_descripcion=req_data.material_descripcion,
        cantidad=req_data.cantidad,
        unidad_medida=req_data.unidad_medida,
        comentarios=req_data.comentarios,
        requerido_por=req_data.requerido_por,
        area=req_data.area,
        proyecto=req_data.proyecto,
        proceso=req_data.proceso,
        main_part=req_data.main_part,
        part_number=req_data.part_number,
        modelo=req_data.modelo,
        descripcion=req_data.descripcion,
        color=req_data.color,
        cantidad_solicitada=req_data.cantidad_solicitada,
        calidad_solicitada=req_data.calidad_solicitada,
        referencia=req_data.referencia,
        supervisor_aprobador=req_data.supervisor_aprobador,
        recibido_por=req_data.recibido_por,
        autorizacion_ehs=req_data.autorizacion_ehs,
        proveedor_sugerido=req_data.proveedor_sugerido
    )
    
    req_dict = requisicion.model_dump()
    req_dict['fecha'] = req_dict['fecha'].isoformat()
    req_dict['created_at'] = req_dict['created_at'].isoformat()
    req_dict['updated_at'] = req_dict['updated_at'].isoformat()
    if req_dict['fecha_requerida']:
        req_dict['fecha_requerida'] = req_dict['fecha_requerida'].isoformat()
    
    await db.requisiciones.insert_one(req_dict)
    
    # Enviar correo de confirmación al usuario
    email_html = f"""
    <div style="font-family: Arial, sans-serif;">
        <h2 style="color: #009E60;">Requisición Creada - DRE Repair Services</h2>
        <p>Su requisición ha sido creada exitosamente.</p>
        <p><strong>Número de solicitud:</strong> {requisicion.numero_solicitud}</p>
        <p><strong>Tipo:</strong> {requisicion.tipo}</p>
        <p><strong>Departamento:</strong> {requisicion.departamento}</p>
        <p><strong>Prioridad:</strong> {requisicion.prioridad}</p>
        <p><strong>Estado:</strong> Pendiente de cotizaciones</p>
        <p style="margin-top: 20px;">
            <a href="{os.environ.get('FRONTEND_URL', '')}" 
               style="background-color: #009E60; color: white; padding: 10px 20px; text-decoration: none; border-radius: 5px;">
               Ver en el Sistema
            </a>
        </p>
    </div>
    """
    
    # Enviar al usuario y a compras
    if current_user and current_user.email:
        await send_email_notification(current_user.email, f"Requisición {requisicion.numero_solicitud} Creada", email_html)
    
    # Notificar a compras
    notification_email = os.environ.get('NOTIFICATION_EMAIL', 'compras-drerepairservices@hotmail.com')
    await send_email_notification(notification_email, f"Nueva Requisición {requisicion.numero_solicitud}", email_html)
    
    return requisicion

@api_router.get("/requisiciones")
async def get_requisiciones(current_user: User = Depends(get_current_user)):
    query = {}
    
    # Filtrar según rol
    if current_user.roles.admin or current_user.roles.compras:
        # Ver todas (incluidas las públicas con solicitante_id = None)
        query = {}
    elif current_user.roles.supervisor or current_user.roles.gerente or current_user.roles.director or current_user.roles.ceo:
        # Ver las que están asignadas para aprobar
        query = {"supervisor_aprobador_id": current_user.id}
    else:
        # Ver solo las propias (excluye las públicas)
        query = {"solicitante_id": current_user.id}
    
    requisiciones = await db.requisiciones.find(query, {"_id": 0}).to_list(1000)
    for req in requisiciones:
        if isinstance(req.get('fecha'), str):
            req['fecha'] = datetime.fromisoformat(req['fecha'])
        if req.get('fecha_requerida') and isinstance(req['fecha_requerida'], str):
            req['fecha_requerida'] = datetime.fromisoformat(req['fecha_requerida'])
        if isinstance(req.get('created_at'), str):
            req['created_at'] = datetime.fromisoformat(req['created_at'])
        if isinstance(req.get('updated_at'), str):
            req['updated_at'] = datetime.fromisoformat(req['updated_at'])
    
    return requisiciones

@api_router.get("/requisiciones/public/{numero_solicitud}")
async def get_requisicion_public(numero_solicitud: str):
    """Consulta pública de estado de requisición (sin autenticación)"""
    req = await db.requisiciones.find_one({"numero_solicitud": numero_solicitud}, {"_id": 0})
    if not req:
        raise HTTPException(status_code=404, detail="Requisición no encontrada")
    
    # Convertir fechas
    if isinstance(req.get('fecha'), str):
        req['fecha'] = datetime.fromisoformat(req['fecha'])
    if req.get('fecha_requerida') and isinstance(req['fecha_requerida'], str):
        req['fecha_requerida'] = datetime.fromisoformat(req['fecha_requerida'])
    if isinstance(req.get('created_at'), str):
        req['created_at'] = datetime.fromisoformat(req['created_at'])
    if isinstance(req.get('updated_at'), str):
        req['updated_at'] = datetime.fromisoformat(req['updated_at'])
    
    # No mostrar montos a usuarios no autenticados
    return {
        "numero_solicitud": req['numero_solicitud'],
        "tipo": req['tipo'],
        "estado": req['estado'],
        "fecha": req['fecha'],
        "departamento": req['departamento']
    }

@api_router.get("/requisiciones/{requisicion_id}")
async def get_requisicion(requisicion_id: str, current_user: User = Depends(get_current_user)):
    req = await db.requisiciones.find_one({"id": requisicion_id}, {"_id": 0})
    if not req:
        raise HTTPException(status_code=404, detail="Requisición no encontrada")
    
    # Convertir fechas
    if isinstance(req.get('fecha'), str):
        req['fecha'] = datetime.fromisoformat(req['fecha'])
    if req.get('fecha_requerida') and isinstance(req['fecha_requerida'], str):
        req['fecha_requerida'] = datetime.fromisoformat(req['fecha_requerida'])
    if isinstance(req.get('created_at'), str):
        req['created_at'] = datetime.fromisoformat(req['created_at'])
    if isinstance(req.get('updated_at'), str):
        req['updated_at'] = datetime.fromisoformat(req['updated_at'])
    
    return req

@api_router.patch("/requisiciones/{requisicion_id}")
async def update_requisicion(requisicion_id: str, updates: Dict[str, Any], current_user: User = Depends(get_current_user)):
    # Solo admin y compras pueden editar
    if not (current_user.roles.admin or current_user.roles.compras):
        raise HTTPException(status_code=403, detail="No tiene permisos para editar")
    
    updates['updated_at'] = datetime.now(timezone.utc).isoformat()
    await db.requisiciones.update_one({"id": requisicion_id}, {"$set": updates})
    return {"message": "Requisición actualizada"}

# ==================== COTIZACIONES ENDPOINTS ====================

@api_router.post("/cotizaciones")
async def create_cotizacion(cotizacion_data: CotizacionCreate, current_user: User = Depends(get_current_user)):
    # Solo compras puede cargar cotizaciones
    if not (current_user.roles.compras or current_user.roles.admin):
        raise HTTPException(status_code=403, detail="No tiene permisos para cargar cotizaciones")
    
    cotizacion = Cotizacion(
        **cotizacion_data.model_dump(),
        cargado_por=current_user.nombre,
        cargado_por_id=current_user.id
    )
    
    cot_dict = cotizacion.model_dump()
    cot_dict['fecha_carga'] = cot_dict['fecha_carga'].isoformat()
    
    await db.cotizaciones.insert_one(cot_dict)
    
    # Actualizar contador de cotizaciones en requisición
    count = await db.cotizaciones.count_documents({"requisicion_id": cotizacion_data.requisicion_id})
    await db.requisiciones.update_one(
        {"id": cotizacion_data.requisicion_id},
        {"$set": {"cotizaciones_count": count, "estado": "cotizaciones_cargadas"}}
    )
    
    return cotizacion

@api_router.get("/cotizaciones/requisicion/{requisicion_id}")
async def get_cotizaciones_by_requisicion(requisicion_id: str, current_user: User = Depends(get_current_user)):
    cotizaciones = await db.cotizaciones.find({"requisicion_id": requisicion_id}, {"_id": 0}).to_list(100)
    for cot in cotizaciones:
        if isinstance(cot.get('fecha_carga'), str):
            cot['fecha_carga'] = datetime.fromisoformat(cot['fecha_carga'])
    return cotizaciones

@api_router.post("/cotizaciones/enviar-aprobacion/{requisicion_id}")
async def enviar_a_aprobacion(requisicion_id: str, current_user: User = Depends(get_current_user)):
    # Solo compras puede enviar a aprobación
    if not (current_user.roles.compras or current_user.roles.admin):
        raise HTTPException(status_code=403, detail="No tiene permisos")
    
    # Verificar que hay al menos 3 cotizaciones
    count = await db.cotizaciones.count_documents({"requisicion_id": requisicion_id})
    if count < 3:
        raise HTTPException(status_code=400, detail="Se requieren al menos 3 cotizaciones")
    
    # Obtener requisición
    req = await db.requisiciones.find_one({"id": requisicion_id})
    if not req:
        raise HTTPException(status_code=404, detail="Requisición no encontrada")
    
    # Obtener cotizaciones
    cotizaciones = await db.cotizaciones.find({"requisicion_id": requisicion_id}, {"_id": 0}).to_list(100)
    
    # Actualizar estado
    await db.requisiciones.update_one(
        {"id": requisicion_id},
        {"$set": {"estado": "enviado_aprobacion"}}
    )
    
    # Enviar correo al supervisor
    supervisor_email = os.environ.get('SUPERVISOR_EMAIL', 'compras-drerepairservices@hotmail.com')
    
    cotizaciones_html = ""
    for i, cot in enumerate(cotizaciones, 1):
        cotizaciones_html += f"""
        <div style="margin: 10px 0; padding: 10px; border: 1px solid #ddd; border-radius: 5px;">
            <h4>Cotización {i} - {cot['proveedor']}</h4>
            <p><strong>Número:</strong> {cot['numero_cotizacion']}</p>
            <p><strong>Monto RD$:</strong> ${cot['monto_rdp']:,.2f}</p>
            <p><strong>Monto US$:</strong> ${cot['monto_usd']:,.2f}</p>
            <p><strong>Descripción:</strong> {cot['descripcion']}</p>
        </div>
        """
    
    email_html = f"""
    <div style="font-family: Arial, sans-serif;">
        <h2 style="color: #009E60;">Requisición Lista para Aprobación - DRE Repair Services</h2>
        <p>La requisición <strong>{req['numero_solicitud']}</strong> tiene {count} cotizaciones cargadas y está lista para su aprobación.</p>
        
        <h3>Detalles de la Requisición:</h3>
        <p><strong>Tipo:</strong> {req['tipo']}</p>
        <p><strong>Departamento:</strong> {req['departamento']}</p>
        <p><strong>Solicitante:</strong> {req['solicitante']}</p>
        <p><strong>Descripción:</strong> {req.get('material_descripcion', req.get('descripcion', 'N/A'))}</p>
        
        <h3>Cotizaciones:</h3>
        {cotizaciones_html}
        
        <p style="margin-top: 20px;">
            <a href="{os.environ.get('FRONTEND_URL', os.environ.get('REACT_APP_BACKEND_URL', ''))}" 
               style="background-color: #009E60; color: white; padding: 10px 20px; text-decoration: none; border-radius: 5px;">
               Ver Solicitud en el Sistema
            </a>
        </p>
    </div>
    """
    
    await send_email_notification(
        supervisor_email,
        f"Aprobación Requerida - Requisición {req['numero_solicitud']}",
        email_html
    )
    
    return {"message": "Enviado a aprobación exitosamente"}

@api_router.post("/cotizaciones/aprobar/{cotizacion_id}")
async def aprobar_cotizacion(cotizacion_id: str, current_user: User = Depends(get_current_user)):
    # Solo supervisores, gerentes, directores y CEO pueden aprobar
    if not (current_user.roles.supervisor or current_user.roles.gerente or 
            current_user.roles.director or current_user.roles.ceo or current_user.roles.admin):
        raise HTTPException(status_code=403, detail="No tiene permisos para aprobar")
    
    cot = await db.cotizaciones.find_one({"id": cotizacion_id})
    if not cot:
        raise HTTPException(status_code=404, detail="Cotización no encontrada")
    
    # Aprobar cotización
    await db.cotizaciones.update_one({"id": cotizacion_id}, {"$set": {"aprobada": True}})
    
    # Actualizar estado de requisición
    await db.requisiciones.update_one(
        {"id": cot['requisicion_id']},
        {"$set": {"estado": "aprobado", "aprobado_por": current_user.nombre}}
    )
    
    return {"message": "Cotización aprobada"}

@api_router.post("/cotizaciones/rechazar/{cotizacion_id}")
async def rechazar_cotizacion(cotizacion_id: str, comentario: str = Body(...), current_user: User = Depends(get_current_user)):
    if not (current_user.roles.supervisor or current_user.roles.gerente or 
            current_user.roles.director or current_user.roles.ceo or current_user.roles.admin):
        raise HTTPException(status_code=403, detail="No tiene permisos para rechazar")
    
    cot = await db.cotizaciones.find_one({"id": cotizacion_id})
    if not cot:
        raise HTTPException(status_code=404, detail="Cotización no encontrada")
    
    # Rechazar cotización
    await db.cotizaciones.update_one(
        {"id": cotizacion_id},
        {"$set": {"rechazada": True, "comentario_rechazo": comentario}}
    )
    
    # Actualizar estado de requisición
    await db.requisiciones.update_one(
        {"id": cot['requisicion_id']},
        {"$set": {"estado": "rechazado", "comentario_rechazo": comentario}}
    )
    
    return {"message": "Cotización rechazada"}

# ==================== PURCHASE DOCUMENT ENDPOINTS ====================

@api_router.post("/purchase-documents")
async def create_pd(pd_data: PDCreate, current_user: User = Depends(get_current_user)):
    # Solo compras y admin pueden crear PDs
    if not (current_user.roles.compras or current_user.roles.admin):
        raise HTTPException(status_code=403, detail="No tiene permisos")
    
    # Calcular campos
    pd_dict = pd_data.model_dump()
    pd_dict['po_date'] = datetime.fromisoformat(pd_data.po_date)
    pd_dict['net_price_rdp'] = pd_dict['unit_price_rdp'] * pd_dict['qty'] + pd_dict['itbis_rdp']
    
    pd_dict = calculate_pd_fields(pd_dict)
    
    pd = PurchaseDocument(**pd_dict)
    
    pd_save = pd.model_dump()
    pd_save['po_date'] = pd_save['po_date'].isoformat()
    pd_save['created_at'] = pd_save['created_at'].isoformat()
    pd_save['updated_at'] = pd_save['updated_at'].isoformat()
    
    await db.purchase_documents.insert_one(pd_save)
    
    # Enviar notificación por email
    req = await db.requisiciones.find_one({"id": pd_data.requisicion_id})
    
    email_html = f"""
    <div style="font-family: Arial, sans-serif;">
        <h2 style="color: #009E60;">Purchase Document Creado - DRE Repair Services</h2>
        <p>Se ha generado un nuevo Purchase Document.</p>
        
        <h3>Información del PD:</h3>
        <table style="border-collapse: collapse; width: 100%;">
            <tr><td style="padding: 8px; border: 1px solid #ddd;"><strong>ID:</strong></td><td style="padding: 8px; border: 1px solid #ddd;">{pd.id}</td></tr>
            <tr><td style="padding: 8px; border: 1px solid #ddd;"><strong>Requisición:</strong></td><td style="padding: 8px; border: 1px solid #ddd;">{req.get('numero_solicitud', 'N/A') if req else 'N/A'}</td></tr>
            <tr><td style="padding: 8px; border: 1px solid #ddd;"><strong>Supplier:</strong></td><td style="padding: 8px; border: 1px solid #ddd;">{pd.supplier}</td></tr>
            <tr><td style="padding: 8px; border: 1px solid #ddd;"><strong>PO:</strong></td><td style="padding: 8px; border: 1px solid #ddd;">{pd.po}</td></tr>
            <tr><td style="padding: 8px; border: 1px solid #ddd;"><strong>Department:</strong></td><td style="padding: 8px; border: 1px solid #ddd;">{pd.department}</td></tr>
            <tr><td style="padding: 8px; border: 1px solid #ddd;"><strong>Description:</strong></td><td style="padding: 8px; border: 1px solid #ddd;">{pd.description}</td></tr>
            <tr><td style="padding: 8px; border: 1px solid #ddd;"><strong>QTY:</strong></td><td style="padding: 8px; border: 1px solid #ddd;">{pd.qty}</td></tr>
            <tr><td style="padding: 8px; border: 1px solid #ddd;"><strong>Net Price RD$:</strong></td><td style="padding: 8px; border: 1px solid #ddd;">${pd.net_price_rdp:,.2f}</td></tr>
            <tr><td style="padding: 8px; border: 1px solid #ddd;"><strong>Net Price US$:</strong></td><td style="padding: 8px; border: 1px solid #ddd;">${pd.net_price_usd:,.2f}</td></tr>
            <tr><td style="padding: 8px; border: 1px solid #ddd;"><strong>Order Status:</strong></td><td style="padding: 8px; border: 1px solid #ddd;">{pd.order_status}</td></tr>
        </table>
        
        <p style="margin-top: 20px;">
            <a href="{os.environ.get('FRONTEND_URL', '')}/dashboard/purchase-documents" 
               style="background-color: #009E60; color: white; padding: 10px 20px; text-decoration: none; border-radius: 5px;">
               Ver Purchase Document
            </a>
        </p>
    </div>
    """
    
    # Notificar a compras y al usuario que creó
    notification_email = os.environ.get('NOTIFICATION_EMAIL', 'compras-drerepairservices@hotmail.com')
    await send_email_notification(notification_email, f"Nuevo PD Creado - {pd.id}", email_html)
    
    return pd

@api_router.get("/purchase-documents")
async def get_pds(current_user: User = Depends(get_current_user)):
    pds = await db.purchase_documents.find({}, {"_id": 0}).to_list(1000)
    for pd in pds:
        if isinstance(pd.get('po_date'), str):
            pd['po_date'] = datetime.fromisoformat(pd['po_date'])
        if isinstance(pd.get('created_at'), str):
            pd['created_at'] = datetime.fromisoformat(pd['created_at'])
        if isinstance(pd.get('updated_at'), str):
            pd['updated_at'] = datetime.fromisoformat(pd['updated_at'])
    return pds

@api_router.get("/purchase-documents/{pd_id}")
async def get_pd(pd_id: str, current_user: User = Depends(get_current_user)):
    pd = await db.purchase_documents.find_one({"id": pd_id}, {"_id": 0})
    if not pd:
        raise HTTPException(status_code=404, detail="PD no encontrado")
    
    if isinstance(pd.get('po_date'), str):
        pd['po_date'] = datetime.fromisoformat(pd['po_date'])
    if isinstance(pd.get('created_at'), str):
        pd['created_at'] = datetime.fromisoformat(pd['created_at'])
    if isinstance(pd.get('updated_at'), str):
        pd['updated_at'] = datetime.fromisoformat(pd['updated_at'])
    
    return pd

@api_router.patch("/purchase-documents/{pd_id}")
async def update_pd(pd_id: str, updates: Dict[str, Any], current_user: User = Depends(get_current_user)):
    if not (current_user.roles.compras or current_user.roles.admin):
        raise HTTPException(status_code=403, detail="No tiene permisos")
    
    # Recalcular campos si es necesario
    pd = await db.purchase_documents.find_one({"id": pd_id})
    if not pd:
        raise HTTPException(status_code=404, detail="PD no encontrado")
    
    old_status = pd.get('order_status', '')
    pd.update(updates)
    pd = calculate_pd_fields(pd)
    pd['updated_at'] = datetime.now(timezone.utc).isoformat()
    
    await db.purchase_documents.update_one({"id": pd_id}, {"$set": pd})
    
    # Enviar notificación si cambió el estado
    new_status = pd.get('order_status', '')
    if old_status != new_status:
        email_html = f"""
        <div style="font-family: Arial, sans-serif;">
            <h2 style="color: #009E60;">Actualización de Purchase Document - DRE Repair Services</h2>
            <p>El estado de un Purchase Document ha sido actualizado.</p>
            
            <table style="border-collapse: collapse; width: 100%;">
                <tr><td style="padding: 8px; border: 1px solid #ddd;"><strong>PD ID:</strong></td><td style="padding: 8px; border: 1px solid #ddd;">{pd['id']}</td></tr>
                <tr><td style="padding: 8px; border: 1px solid #ddd;"><strong>Supplier:</strong></td><td style="padding: 8px; border: 1px solid #ddd;">{pd['supplier']}</td></tr>
                <tr><td style="padding: 8px; border: 1px solid #ddd;"><strong>PO:</strong></td><td style="padding: 8px; border: 1px solid #ddd;">{pd['po']}</td></tr>
                <tr><td style="padding: 8px; border: 1px solid #ddd;"><strong>Estado Anterior:</strong></td><td style="padding: 8px; border: 1px solid #ddd;">{old_status}</td></tr>
                <tr><td style="padding: 8px; border: 1px solid #ddd;"><strong>Estado Nuevo:</strong></td><td style="padding: 8px; border: 1px solid #ddd; font-weight: bold; color: #009E60;">{new_status}</td></tr>
                <tr><td style="padding: 8px; border: 1px solid #ddd;"><strong>QTY Pending:</strong></td><td style="padding: 8px; border: 1px solid #ddd;">{pd.get('qty_pending', 0)}</td></tr>
            </table>
            
            <p style="margin-top: 20px;">
                <a href="{os.environ.get('FRONTEND_URL', '')}/dashboard/purchase-documents" 
                   style="background-color: #009E60; color: white; padding: 10px 20px; text-decoration: none; border-radius: 5px;">
                   Ver Purchase Document
                </a>
            </p>
        </div>
        """
        
        notification_email = os.environ.get('NOTIFICATION_EMAIL', 'compras-drerepairservices@hotmail.com')
        await send_email_notification(notification_email, f"Actualización PD - {pd['id']}", email_html)
    
    return {"message": "PD actualizado"}

@api_router.delete("/purchase-documents/{pd_id}")
async def delete_pd(pd_id: str, current_user: User = Depends(get_current_user)):
    if not current_user.roles.admin:
        raise HTTPException(status_code=403, detail="Solo administradores pueden eliminar")
    
    await db.purchase_documents.delete_one({"id": pd_id})
    return {"message": "PD eliminado"}

# ==================== UTILITIES ENDPOINTS ====================

@api_router.get("/exchange-rate")
async def get_current_exchange_rate():
    rate = await get_exchange_rate()
    return {"rate": rate, "currency": "DOP"}

@api_router.get("/users")
async def get_users(current_user: User = Depends(get_current_user)):
    if not current_user.roles.admin:
        raise HTTPException(status_code=403, detail="No tiene permisos")
    
    users = await db.users.find({}, {"_id": 0, "password": 0}).to_list(1000)
    return users

@api_router.patch("/users/{user_id}")
async def update_user(user_id: str, updates: Dict[str, Any], current_user: User = Depends(get_current_user)):
    if not current_user.roles.admin:
        raise HTTPException(status_code=403, detail="No tiene permisos")
    
    await db.users.update_one({"id": user_id}, {"$set": updates})
    return {"message": "Usuario actualizado"}

@api_router.delete("/users/{user_id}")
async def delete_user(user_id: str, current_user: User = Depends(get_current_user)):
    if not current_user.roles.admin:
        raise HTTPException(status_code=403, detail="Solo administradores pueden eliminar usuarios")
    
    await db.users.delete_one({"id": user_id})
    return {"message": "Usuario eliminado"}

@api_router.get("/export/requisiciones")
async def export_requisiciones(current_user: User = Depends(get_current_user)):
    """Exporta requisiciones a Excel"""
    requisiciones = await db.requisiciones.find({}, {"_id": 0}).to_list(10000)
    
    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Requisiciones"
    
    # Headers
    headers = ["Número", "Tipo", "Fecha", "Solicitante", "Departamento", "Estado", 
               "Prioridad", "Descripción", "Supervisor Aprobador"]
    ws.append(headers)
    
    # Estilo de headers
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="009E60", end_color="009E60", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Datos
    for req in requisiciones:
        fecha = req['fecha'] if isinstance(req['fecha'], str) else req['fecha'].strftime('%Y-%m-%d')
        ws.append([
            req['numero_solicitud'],
            req['tipo'],
            fecha,
            req['solicitante'],
            req['departamento'],
            req['estado'],
            req['prioridad'],
            req.get('material_descripcion', req.get('descripcion', '')),
            req['supervisor_aprobador']
        ])
    
    # Guardar en memoria
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=requisiciones.xlsx"}
    )

@api_router.get("/export/purchase-documents")
async def export_pds(current_user: User = Depends(get_current_user)):
    """Exporta Purchase Documents a Excel"""
    pds = await db.purchase_documents.find({}, {"_id": 0}).to_list(10000)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Purchase Documents"
    
    headers = ["ID", "Part Numbers", "Supplier", "PO", "PO Date", "Department", "Description", 
               "QTY", "Unit Price RD$", "Net Price RD$", "Unit Price US$", "Net Price US$", 
               "QTY Received", "QTY Pending", "Order Status"]
    ws.append(headers)
    
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="009E60", end_color="009E60", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    for pd in pds:
        po_date = pd['po_date'] if isinstance(pd['po_date'], str) else pd['po_date'].strftime('%Y-%m-%d')
        ws.append([
            pd['id'],
            pd['part_numbers'],
            pd['supplier'],
            pd['po'],
            po_date,
            pd['department'],
            pd['description'],
            pd['qty'],
            pd['unit_price_rdp'],
            pd['net_price_rdp'],
            pd['unit_price_usd'],
            pd['net_price_usd'],
            pd['qty_received'],
            pd['qty_pending'],
            pd['order_status']
        ])
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=purchase_documents.xlsx"}
    )

# Upload image endpoint
@api_router.post("/upload/image")
async def upload_image(file: UploadFile = File(...)):
    """Sube imagen y retorna la URL"""
    file_ext = file.filename.split('.')[-1]
    file_name = f"{uuid.uuid4()}.{file_ext}"
    file_path = UPLOAD_DIR / file_name
    
    async with aiofiles.open(file_path, 'wb') as f:
        content = await file.read()
        await f.write(content)
    
    return {"url": f"/uploads/{file_name}"}

# ==================== MAIN ====================

app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()